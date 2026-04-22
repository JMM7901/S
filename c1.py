import os
import cv2
import json
import time
import shutil
import traceback
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from PIL import Image
from datetime import datetime
import google.generativeai as genai
import pytesseract
import matplotlib.pyplot as plt

# ==========================================
# 0. DEBUG MODE TOGGLE
# ==========================================
# Set to True to see images inline in Databricks. Set to False for fast production runs.
DEBUG_MODE = True  

def show_debug_image(title, img):
    if not DEBUG_MODE: return
    display_img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB) if len(img.shape) == 3 else img
    plt.figure(figsize=(15, 20)) 
    plt.imshow(display_img, cmap='gray' if len(img.shape)==2 else None)
    plt.title(title, fontsize=20, color='red')
    plt.axis('off')
    plt.show()

def show_alignment_overlay(aligned_img, template_img):
    if not DEBUG_MODE: return
    overlay = cv2.addWeighted(aligned_img, 0.5, template_img, 0.5, 0)
    show_debug_image("GHOST OVERLAY: Aligned Target (50%) + Blank Template (50%)", overlay)

# ==========================================
# 1. CONFIGURATION & PATHS
# ==========================================
API_KEY = "YOUR_GEMINI_API_KEY"
genai.configure(api_key=API_KEY)

PRICE_INPUT = 0.075  
PRICE_OUTPUT = 0.30  

TARGET_WIDTH = 2550
TARGET_HEIGHT = 3300

# Databricks Paths
LOCAL_WORKSPACE_DIR = "/tmp/medical_claims_processing/"
VOLUME_DIR = "/Volumes/your_catalog/your_schema/your_volume/"

INPUT_FOLDER = os.path.join(VOLUME_DIR, "input_images")
SAMPLE_CMS1500 = os.path.join(VOLUME_DIR, "templates", "sample_cms1500.jpg")
SAMPLE_CMS1450 = os.path.join(VOLUME_DIR, "templates", "sample_cms1450.jpg")

LOCAL_EXCEL_1500 = os.path.join(LOCAL_WORKSPACE_DIR, "output_cms1500.xlsx")
VOLUME_EXCEL_1500 = os.path.join(VOLUME_DIR, "output_cms1500.xlsx")

LOCAL_EXCEL_1450 = os.path.join(LOCAL_WORKSPACE_DIR, "output_cms1450.xlsx")
VOLUME_EXCEL_1450 = os.path.join(VOLUME_DIR, "output_cms1450.xlsx")

LOCAL_INVENTORY = os.path.join(LOCAL_WORKSPACE_DIR, "inventory.xlsx")
VOLUME_INVENTORY = os.path.join(VOLUME_DIR, "inventory.xlsx")
LOCAL_TOKENS = os.path.join(LOCAL_WORKSPACE_DIR, "token_calculation.xlsx")
VOLUME_TOKENS = os.path.join(VOLUME_DIR, "token_calculation.xlsx")
PROCESS_LOG_PATH = os.path.join(VOLUME_DIR, "process_log.txt")

os.makedirs(LOCAL_WORKSPACE_DIR, exist_ok=True)

# ==========================================
# 2. COORDINATE MAP (2-Point Bounding Box)
# ==========================================
# IMPORTANT: Map these using the 2550x3300 templates you generated.
COORDINATE_MAP = {
    "CMS-1500": {
        "Claim_Form_Number": {"x1": 2000, "y1": 50, "x2": 2400, "y2": 150, "type": "text"},
        "1_LOB_Medicare": {"x1": 100, "y1": 250, "x2": 150, "y2": 300, "type": "checkbox"},
        # Add the rest of your x1,y1 to x2,y2 coordinates here...
    },
    "CMS-1450": {
        "Claim_Form_Number": {"x1": 2000, "y1": 50, "x2": 2400, "y2": 150, "type": "text"},
        # Add UB-04 coordinates here...
    }
}

# ==========================================
# 3. LOGGING, UTILITIES & EXCEL FORMATTING
# ==========================================
def log_process_status(filename, status, reason=""):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_entry = f"[{timestamp}] {filename} : {status}" + (f" - {reason}" if reason else "")
    print(log_entry)
    with open(PROCESS_LOG_PATH, "a") as f:
        f.write(log_entry + "\n")

def sync_to_volume(local_path, volume_path):
    try:
        if os.path.exists(volume_path): os.remove(volume_path) 
        if os.path.exists(local_path): shutil.copy2(local_path, volume_path) 
    except Exception as e:
        print(f"Volume Sync Error: {e}")

def save_and_format_excel(df, local_path, volume_path, table_name="DataTable"):
    df.to_excel(local_path, index=False, engine='openpyxl')
    wb = openpyxl.load_workbook(local_path)
    ws = wb.active

    if len(df) > 0:
        max_col, max_row = len(df.columns), len(df) + 1 
        tab = Table(displayName=table_name.replace("-", "_"), ref=f"A1:{get_column_letter(max_col)}{max_row}")
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        ws.add_table(tab)

    for col_idx, column in enumerate(ws.columns, 1):
        col_letter = get_column_letter(col_idx)
        max_length = max([len(str(cell.value)) for cell in column if cell.value] + [0])
        for cell in column:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        ws.column_dimensions[col_letter].width = min(max_length + 2, 45)

    wb.save(local_path)
    sync_to_volume(local_path, volume_path)

def load_or_create_excel(local_path, volume_path, columns, table_name="InitTable"):
    if os.path.exists(volume_path):
        shutil.copy2(volume_path, local_path)
        return pd.read_excel(local_path, engine='openpyxl')
    else:
        df = pd.DataFrame(columns=columns)
        save_and_format_excel(df, local_path, volume_path, table_name)
        return df

def get_or_create_inventory():
    if os.path.exists(VOLUME_INVENTORY):
        shutil.copy2(VOLUME_INVENTORY, LOCAL_INVENTORY)
        return pd.read_excel(LOCAL_INVENTORY, engine='openpyxl')
    else:
        valid_exts = ('.png', '.jpg', '.jpeg', '.tiff', '.tif', '.bmp', '.webp')
        files = sorted([f for f in os.listdir(INPUT_FOLDER) if f.lower().endswith(valid_exts)])
        df = pd.DataFrame([{"Index": idx, "File_Name": f, "Status": "PENDING"} for idx, f in enumerate(files)])
        save_and_format_excel(df, LOCAL_INVENTORY, VOLUME_INVENTORY, table_name="Inventory")
        return df

def log_token_usage_excel(filename, in_tokens, out_tokens, status="SUCCESS"):
    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    total_cost = ((in_tokens / 1_000_000) * PRICE_INPUT) + ((out_tokens / 1_000_000) * PRICE_OUTPUT)
    df = load_or_create_excel(LOCAL_TOKENS, VOLUME_TOKENS, ['filename', 'input_tokens', 'output_tokens', 'Total_Cost', 'processed_flag', 'processing_date'], "TokenTable")
    mask = df['filename'] == filename
    if mask.any():
        df.loc[mask, ['input_tokens', 'output_tokens', 'Total_Cost', 'processed_flag', 'processing_date']] = [in_tokens, out_tokens, round(total_cost, 6), status, current_time]
    else:
        df = pd.concat([df, pd.DataFrame([{'filename': filename, 'input_tokens': in_tokens, 'output_tokens': out_tokens, 'Total_Cost': round(total_cost, 6), 'processed_flag': status, 'processing_date': current_time}])], ignore_index=True)
    save_and_format_excel(df, LOCAL_TOKENS, VOLUME_TOKENS, "TokenTable")

# ==========================================
# 4. COMPUTER VISION PIPELINE
# ==========================================
def deskew_image(pil_image):
    if pil_image is None: return None
    
    if pil_image.mode in ("RGBA", "P"):
        pil_image = pil_image.convert("RGB")
        
    img = np.array(pil_image)
    img = cv2.cvtColor(img, cv2.COLOR_RGB2BGR)
    
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    thresh = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV | cv2.THRESH_OTSU)[1]
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (5, 5))
    dilated = cv2.dilate(thresh, kernel, iterations=2)
    contours, _ = cv2.findContours(dilated, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    aspect_ratios = [h/w for cnt in contours for x, y, w, h in [cv2.boundingRect(cnt)] if w > 20 and h > 20]
    median_ratio = np.median(aspect_ratios) if aspect_ratios else 0.5 
    
    if median_ratio > 1.2:
        img = cv2.rotate(img, cv2.ROTATE_90_CLOCKWISE)
        thresh = cv2.rotate(thresh, cv2.ROTATE_90_CLOCKWISE) 

    h, w = img.shape[:2]
    scores = []
    angles = np.arange(-5, 5.1, 0.5) 
    for angle in angles:
        M = cv2.getRotationMatrix2D((w//2, h//2), angle, 1.0)
        rotated_thresh = cv2.warpAffine(thresh, M, (w, h), flags=cv2.INTER_NEAREST)
        hist = np.sum(rotated_thresh, axis=1)
        scores.append(np.sum((hist[1:] - hist[:-1]) ** 2))

    best_micro_angle = angles[np.argmax(scores)]
    if abs(best_micro_angle) > 0.1:
        M = cv2.getRotationMatrix2D((w // 2, h // 2), best_micro_angle, 1.0)
        img = cv2.warpAffine(img, M, (w, h), flags=cv2.INTER_CUBIC, borderMode=cv2.BORDER_CONSTANT, borderValue=(255, 255, 255))
    return img


def order_points(pts):
    """Sorts corners into: Top-Left, Top-Right, Bottom-Right, Bottom-Left"""
    rect = np.zeros((4, 2), dtype="float32")
    s = pts.sum(axis=1)
    rect[0] = pts[np.argmin(s)] # Top-Left has smallest x+y
    rect[2] = pts[np.argmax(s)] # Bottom-Right has largest x+y
    
    diff = np.diff(pts, axis=1)
    rect[1] = pts[np.argmin(diff)] # Top-Right has smallest x-y
    rect[3] = pts[np.argmax(diff)] # Bottom-Left has largest x-y
    return rect

def get_document_corners(image, debug_name="Image"):
    """Isolates horizontal and vertical lines to find the form's grid bounding box."""
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    
    # 1. Binarize the image (Black background, White lines/text)
    thresh = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV | cv2.THRESH_OTSU)[1]

    # 2. Define the minimum length a line must be to be considered part of the grid
    # For a 2550x3300 image, a grid line should be at least ~50 pixels long
    line_min_length = max(image.shape[1], image.shape[0]) // 50

    # 3. Extract purely VERTICAL lines
    vertical_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, line_min_length))
    vertical_lines = cv2.morphologyEx(thresh, cv2.MORPH_OPEN, vertical_kernel, iterations=2)

    # 4. Extract purely HORIZONTAL lines
    horizontal_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (line_min_length, 1))
    horizontal_lines = cv2.morphologyEx(thresh, cv2.MORPH_OPEN, horizontal_kernel, iterations=2)

    # 5. Merge them into a single "Skeleton Grid"
    grid_mask = cv2.addWeighted(vertical_lines, 0.5, horizontal_lines, 0.5, 0.0)
    grid_mask = cv2.threshold(grid_mask, 50, 255, cv2.THRESH_BINARY)[1]

    # Dilate the grid slightly to connect any dashed lines or broken intersections
    join_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (15, 15))
    grid_mask = cv2.dilate(grid_mask, join_kernel, iterations=2)

    # --- DEBUG: Show the Grid Skeleton ---
    if DEBUG_MODE:
        show_debug_image(f"DEBUG: Skeleton Grid for {debug_name}", grid_mask)
    # -------------------------------------

    # 6. Find the largest shape on this clean grid
    cnts, _ = cv2.findContours(grid_mask, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    if not cnts:
        raise ValueError(f"Could not find any grid lines in {debug_name}.")
        
    # Sort by area and grab the absolute largest contour (The Outer Box)
    largest_contour = max(cnts, key=cv2.contourArea)

    # 7. Get the Rotated Bounding Box (Mathematically guarantees 4 extreme corners)
    rect = cv2.minAreaRect(largest_contour)
    box = cv2.boxPoints(rect)
    box = np.float32(box) # Convert to standard floats for OpenCV

    # --- DEBUG: Show the Final Found Box ---
    if DEBUG_MODE:
        debug_canvas = image.copy()
        cv2.drawContours(debug_canvas, [np.int32(box)], 0, (0, 255, 0), 10)
        show_debug_image(f"Found Outer Bounds: {debug_name}", debug_canvas)
    # ---------------------------------------

    # Order the points: Top-Left, Top-Right, Bottom-Right, Bottom-Left
    return order_points(box)

def preprocess_and_align(target_path, template_path):
    # 1. Load Images
    pil_img = Image.open(target_path)
    deskewed_cv_img = deskew_image(pil_img)
    target_img = standardize_image_size(deskewed_cv_img)
    
    template_img_raw = cv2.imread(template_path, cv2.IMREAD_COLOR)
    template_img = standardize_image_size(template_img_raw)

    # 2. Find the 4 corners of the outer box in BOTH images
    try:
        target_corners = get_document_corners(target_img, "Filled Target Form")
        template_corners = get_document_corners(template_img, "Blank Template")
    except ValueError as e:
        # Fallback: If it's a terrible scan and it can't find the box, 
        # just return the standardized image and hope the coordinates hit.
        print(f"Warning: {e} Falling back to raw resize.")
        return target_img 

    # 3. Calculate the Perspective Transform Matrix (The CamScanner Math)
    # Instead of Homography with 1,000 messy points, we use exactly 4 perfect points.
    transform_matrix = cv2.getPerspectiveTransform(target_corners, template_corners)

    # 4. Warp the Target to perfectly match the Template
    aligned_img = cv2.warpPerspective(target_img, transform_matrix, (TARGET_WIDTH, TARGET_HEIGHT))
    
    show_alignment_overlay(aligned_img, template_img)
    return aligned_img


def standardize_image_size(img):
    return cv2.resize(img, (TARGET_WIDTH, TARGET_HEIGHT), interpolation=cv2.INTER_CUBIC)

def preprocess_and_align(target_path, template_path, max_features=5000, match_percent=0.15):
    pil_img = Image.open(target_path)
    deskewed_cv_img = deskew_image(pil_img)
    show_debug_image("STEP 1: Deskewed Raw Image", deskewed_cv_img)
    
    template_img_raw = cv2.imread(template_path, cv2.IMREAD_COLOR)

    target_img = standardize_image_size(deskewed_cv_img)
    template_img = standardize_image_size(template_img_raw)
    show_debug_image("STEP 2: Forced Aspect Ratio (2550 x 3300)", target_img)

    target_gray = cv2.cvtColor(target_img, cv2.COLOR_BGR2GRAY)
    template_gray = cv2.cvtColor(template_img, cv2.COLOR_BGR2GRAY)

    orb = cv2.ORB_create(max_features)
    keypoints1, descriptors1 = orb.detectAndCompute(target_gray, None)
    keypoints2, descriptors2 = orb.detectAndCompute(template_gray, None)

    matcher = cv2.DescriptorMatcher_create(cv2.DESCRIPTOR_MATCHER_BRUTEFORCE_HAMMING)
    matches = sorted(matcher.match(descriptors1, descriptors2, None), key=lambda x: x.distance)
    matches = matches[:int(len(matches) * match_percent)]

    if len(matches) < 50: raise ValueError("Alignment Failed: Not enough matching features.")

    points1 = np.float32([keypoints1[m.queryIdx].pt for m in matches])
    points2 = np.float32([keypoints2[m.trainIdx].pt for m in matches])

    h_matrix, _ = cv2.findHomography(points1, points2, cv2.RANSAC)
    aligned_img = cv2.warpPerspective(target_img, h_matrix, (TARGET_WIDTH, TARGET_HEIGHT))
    
    show_alignment_overlay(aligned_img, template_img)
    return aligned_img

def identify_form_type(aligned_img):
    h, w = aligned_img.shape[:2]
    top_right_crop = aligned_img[0:int(h*0.1), int(w*0.7):w] 
    
    extracted_text = pytesseract.image_to_string(top_right_crop, config='--psm 6').strip()
    
    if "1500" in extracted_text: return "CMS-1500"
    elif "UB" in extracted_text or "1450" in extracted_text: return "CMS-1450"
    else: raise ValueError(f"Could not identify form. Found: '{extracted_text}'")

def extract_zonal_data(aligned_img, form_type):
    coords = COORDINATE_MAP.get(form_type)
    if not coords: raise ValueError(f"No coordinate map for {form_type}")

    # --- DEBUG VISUALIZATION ---
    if DEBUG_MODE:
        debug_canvas = aligned_img.copy()
        for field_name, box in coords.items():
            # Draw blue boxes for LLM, red for Tesseract
            color = (255, 0, 0) if box['type'] == 'llm' else (0, 0, 255)
            cv2.rectangle(debug_canvas, (box['x1'], box['y1']), (box['x2'], box['y2']), color, 4)
        show_debug_image("STEP 3: Cropping Zones (Blue=LLM, Red=Tesseract)", debug_canvas)
    # ---------------------------

    raw_data = {}
    vision_in_tokens = 0
    vision_out_tokens = 0
    model = genai.GenerativeModel('gemini-1.5-flash') # Initialize vision model once

    for field_name, box in coords.items():
        x1, y1, x2, y2 = box['x1'], box['y1'], box['x2'], box['y2']
        w, h = x2 - x1, y2 - y1
        pad_x, pad_y = int(w * 0.05), int(h * 0.05)
        
        crop = aligned_img[max(0, y1-pad_y) : y2+pad_y, max(0, x1-pad_x) : x2+pad_x]
        
        # ROUTE 1: Fast & Free Tesseract
        if box['type'] == 'text':
            raw_data[field_name] = pytesseract.image_to_string(crop, config='--psm 6').strip()
            
        # ROUTE 2: Advanced LLM Vision
        elif box['type'] == 'llm':
            # Convert OpenCV (BGR) to PIL (RGB) for Gemini
            rgb_img = cv2.cvtColor(crop, cv2.COLOR_BGR2RGB)
            pil_img = Image.fromarray(rgb_img)
            
            # Send the cropped image + the specific prompt
            response = model.generate_content([box['prompt'], pil_img])
            raw_data[field_name] = response.text.strip()
            
            # Track Tokens (Optional but recommended for cost tracking)
            vision_in_tokens += model.count_tokens([box['prompt'], pil_img]).total_tokens
            vision_out_tokens += model.count_tokens(response.text).total_tokens

    return raw_data, vision_in_tokens, vision_out_tokens

# ==========================================
# 5. LLM TEXT CORRECTION
# ==========================================
def correct_data_with_llm(raw_dict, form_type):
    prompt = f"""
    You are an expert medical billing data standardization engine. 
    Clean this raw OCR output from a {form_type} form. Fix 0/O and 1/l swaps. Standardize dates to MM/DD/YYYY. 
    Consolidate check boxes (e.g., if Male is YES and Female is NO, map to Gender: Male).
    Do NOT guess missing data. If blank, return null. Return ONLY raw JSON.

    RAW OCR: {json.dumps(raw_dict)}

    TARGET SCHEMA:
    {{
      "Form_Type": "{form_type}",
      "Claim_Form_Number": "",
      "{form_type.replace('-', '_')}_Data": {{
          "1_LOB": "", "1a_Insured_ID": "", "2_Patient_Name": ""
      }}
    }}
    """
    model = genai.GenerativeModel('gemini-1.5-flash')
    response = model.generate_content(prompt, generation_config=genai.GenerationConfig(response_mime_type="application/json", temperature=0.0))
    return json.loads(response.text), model.count_tokens(prompt).total_tokens, model.count_tokens(response.text).total_tokens


# ==========================================
# 6. MAIN HYBRID PIPELINE
# ==========================================
def process_pipeline(start_index, end_index):
    df_inventory = get_or_create_inventory()
    
    df_1500 = load_or_create_excel(LOCAL_EXCEL_1500, VOLUME_EXCEL_1500, ["File_Name", "Form_Type", "Claim_Form_Number"], "CMS_1500_Data")
    df_1450 = load_or_create_excel(LOCAL_EXCEL_1450, VOLUME_EXCEL_1450, ["File_Name", "Form_Type", "Claim_Form_Number"], "CMS_1450_Data")

    if df_inventory.empty:
        print("No files to process.")
        return

    start_index = max(0, start_index)
    end_index = min(len(df_inventory) - 1, end_index)

    print(f"--- Starting Pipeline: Index {start_index} to {end_index} ---")

    for index in range(start_index, end_index + 1):
        file_name = df_inventory.iloc[index]['File_Name']
        file_path = os.path.join(INPUT_FOLDER, file_name)
        
        if file_name in df_1500["File_Name"].values or file_name in df_1450["File_Name"].values: 
            continue
            
        print(f"\n[{index}] Processing: {file_name}")
        try:
            # 1. Align
            try: aligned_img = preprocess_and_align(file_path, SAMPLE_CMS1500)
            except Exception as e: raise Exception(f"Alignment Phase Failed: {str(e)}")

            # 2. Identify
            try:
                form_type = identify_form_type(aligned_img)
                if form_type == "CMS-1450": aligned_img = preprocess_and_align(file_path, SAMPLE_CMS1450)
            except Exception as e: raise Exception(f"Identification Phase Failed: {str(e)}")

            # 3. Zonal OCR
            try: raw_messy_dict, vis_in, vis_out = extract_zonal_data(aligned_img, form_type)
            except Exception as e: raise Exception(f"Zonal OCR Phase Failed: {str(e)}")

            # 4. LLM Clean
            try: 
                clean_json, clean_in, clean_out = correct_data_with_llm(raw_messy_dict, form_type)
                total_in_tok = vis_in + clean_in
                total_out_tok = vis_out + clean_out
            except Exception as e: raise Exception(f"LLM Formatting Phase Failed: {str(e)}")

            # 5. Routing & Save
            row_data = {"File_Name": file_name, "Form_Type": form_type, "Claim_Form_Number": clean_json.get("Claim_Form_Number", "")}
            row_data.update(clean_json.get(f"{form_type.replace('-', '_')}_Data", {}))
            
            if form_type == "CMS-1500":
                df_1500 = pd.concat([df_1500, pd.DataFrame([row_data])], ignore_index=True)
                save_and_format_excel(df_1500, LOCAL_EXCEL_1500, VOLUME_EXCEL_1500, table_name="CMS_1500_Data")
            
            elif form_type == "CMS-1450":
                df_1450 = pd.concat([df_1450, pd.DataFrame([row_data])], ignore_index=True)
                save_and_format_excel(df_1450, LOCAL_EXCEL_1450, VOLUME_EXCEL_1450, table_name="CMS_1450_Data")

            # 6. Log Success
            df_inventory.loc[index, "Status"] = "SUCCESS"
            save_and_format_excel(df_inventory, LOCAL_INVENTORY, VOLUME_INVENTORY, table_name="Inventory")
            
            log_token_usage_excel(file_name, total_in_tok, total_out_tok, "SUCCESS")
            log_process_status(file_name, "SUCCESS")

        except Exception as e:
            error_reason = str(e)
            print(f"   -> ERROR: {error_reason}")
            log_process_status(file_name, "FAILED", reason=error_reason)
            log_token_usage_excel(file_name, 0, 0, "FAILED")
            
            df_inventory.loc[index, "Status"] = "FAILED"
            save_and_format_excel(df_inventory, LOCAL_INVENTORY, VOLUME_INVENTORY, table_name="Inventory")
            continue

if __name__ == "__main__":
    try:
        dbutils.widgets.text("start_index", "0", "Start Index")
        dbutils.widgets.text("end_index", "100", "End Index")
        s_val = dbutils.widgets.get("start_index")
        e_val = dbutils.widgets.get("end_index")
    except:
        s_val, e_val = "0", "999999"

    start = int(s_val) if s_val.strip() else 0
    end = int(e_val) if e_val.strip() else 999999

    process_pipeline(start_index=start, end_index=end)
