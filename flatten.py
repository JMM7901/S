import os
from datetime import datetime
from zoneinfo import ZoneInfo
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

def flatten_and_save(pass1_data, pass2_data, filename, filetype, attachment_id, user_id):
    try:
        print(f"      -> Pushing data to Excel for {filename}...")

        # 1. Safely handle the LLM output (in case it hallucinated a string)
        if not isinstance(pass1_data, dict):
            pass1_data = {}

        ist_time = datetime.now(ZoneInfo("Asia/Kolkata")).strftime("%Y-%m-%d %H:%M:%S")
        status = "Success" if pass1_data else "Failed/Empty"

        # 2. Define the exact keys from your prompt
        expected_keys = [
            "ONSITE_AUDIT", "ONSITE_PAGE_NUM",
            "DESK_AUDIT", "DESK_PAGE_NUM",
            "BOTH_DESK_ONSITE",
            "PRE_PAYMENT", "PAYMENT_PAGE_NUM",
            "VOLUME_OF_MEDICAL_RECORDS", "VOLUME_PAGE_NUM",
            "POST_PAYMENT", "POST_PAYMENT_PAGE_NUM",
            "CODE_EDITS", "CODE_EDITS_PAGE_NUM",
            "LIMITATION", "LIMITATION_PAGE_NUM"
        ]

        # 3. Build headers
        metadata_headers = [
            "FILE NAME", "FILE EXTENSION", "ATTACHMENT ID", 
            "LLM EXTRACTION STATUS", "RUN TYPE", "LLM USER ID", "LLM PROCESS TIMESTAMP"
        ]
        all_headers = metadata_headers + expected_keys

        # 4. Build the single row of data (using your exact variables)
        row_data = [
            filename, 
            filetype, 
            attachment_id, 
            status, 
            "LLM", 
            user_id, 
            ist_time
        ]

        # Extract the values directly from pass1_data (the flat JSON)
        for key in expected_keys:
            row_data.append(str(pass1_data.get(key, "")))

        # ==========================================
        # 5. EXCEL FILE HANDLING (No Pandas)
        # ==========================================
        
        # Open existing file or create new one if it doesn't exist
        if os.path.exists(MASTER_EXCEL_PATH):
            wb = load_workbook(MASTER_EXCEL_PATH)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(all_headers) # Write headers on row 1

        # Push the new row to the very bottom
        ws.append(row_data)

        # ==========================================
        # 6. DESIGN THE EXCEL (Table & Wrap)
        # ==========================================
        
        max_row = ws.max_row
        max_col = ws.max_column
        col_letter = get_column_letter(max_col)

        # Clear any existing tables so we can stretch it over the new row
        if ws.tables:
            for t in list(ws.tables):
                ws._tables.remove(t)

        # Turn the entire data range into an official Excel Table
        table_ref = f"A1:{col_letter}{max_row}"
        tab = Table(displayName="ContractExtractions", ref=table_ref)
        
        # "TableStyleMedium9" is the standard professional Blue/White striped table
        style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
        tab.tableStyleInfo = style
        ws.add_table(tab)

        # Apply Text Wrapping and align text to the top
        for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')

        # Set a wide column width so the text isn't squished
        for col in range(1, max_col + 1):
            ws.column_dimensions[get_column_letter(col)].width = 35

        # Save the file
        wb.save(MASTER_EXCEL_PATH)
        print(f"      -> Success! Row appended and Excel designed.")

    except Exception as e:
        print(f"      [Error] Failed to push to Excel: {e}")
